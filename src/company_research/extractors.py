from __future__ import annotations

import csv
import io
import logging
import re
import zipfile
from html.parser import HTMLParser
from pathlib import Path
from urllib.parse import urljoin

from .models import ExtractedDocument, FetchResult
from .ranking import classify_source

logger = logging.getLogger(__name__)


def chunk_text(text: str, max_chars: int = 4096) -> list[str]:
    """Normalize text and split it into chunks suitable for storage and retrieval."""
    cleaned = re.sub(r"\s+", " ", text).strip()

    if not cleaned:
        logger.debug("No text available to chunk")
        return []

    chunks: list[str] = []
    start = 0

    while start < len(cleaned):
        end = min(start + max_chars, len(cleaned))

        if end < len(cleaned):
            split = cleaned.rfind(". ", start, end)

            if split > start + 400:
                end = split + 1

        chunks.append(cleaned[start:end].strip())
        start = end

    logger.debug("Chunked text_chars=%s into %s chunks", len(cleaned), len(chunks))

    return chunks


class _HTMLTextExtractor(HTMLParser):
    """Small HTML parser that collects text, title, links, and metadata."""

    def __init__(self, base_url: str):
        """Initialize parser state for one HTML document."""
        super().__init__()

        self.base_url = base_url
        self.title = ""
        self._in_title = False
        self._skip = False
        self.text_parts: list[str] = []
        self.links: list[str] = []
        self.meta: dict[str, str] = {}

    def handle_starttag(self, tag: str, attrs: list[tuple[str, str | None]]) -> None:
        """Collect metadata and links, and mark tags whose content should be skipped."""
        attrs_dict = {key.lower(): value or "" for key, value in attrs}

        if tag in {"script", "style", "noscript"}:
            self._skip = True

        if tag == "title":
            self._in_title = True

        if tag == "a" and attrs_dict.get("href"):
            self.links.append(urljoin(self.base_url, attrs_dict["href"]))

        if tag == "meta":
            name = (attrs_dict.get("name") or attrs_dict.get("property") or "").lower()
            content = attrs_dict.get("content", "")

            if name and content:
                self.meta[name] = content

    def handle_endtag(self, tag: str) -> None:
        """Close parser state and add word boundaries after block-like tags."""
        if tag in {"script", "style", "noscript"}:
            self._skip = False

        if tag == "title":
            self._in_title = False

        if tag in {"p", "div", "section", "article", "tr", "li", "h1", "h2", "h3"}:
            self.text_parts.append(" ")

    def handle_data(self, data: str) -> None:
        """Collect visible text data from the current HTML parser position."""
        if self._skip:
            return

        stripped = data.strip()

        if not stripped:
            return

        if self._in_title:
            self.title += stripped

        self.text_parts.append(stripped)


def extract_html(fetch: FetchResult) -> ExtractedDocument:
    """Extract visible text, links, and metadata from an HTML fetch result."""
    logger.info("Extracting HTML from %s", fetch.final_url)

    html = fetch.body.decode("utf-8", errors="replace")
    parser = _HTMLTextExtractor(fetch.final_url)
    parser.feed(html)

    title = parser.title or parser.meta.get("og:title", "")
    publisher = parser.meta.get("og:site_name", "") or parser.meta.get("author", "")
    text = " ".join(parser.text_parts)
    chunks = chunk_text(text)

    logger.info(
        "Extracted HTML title=%r text_chars=%s chunks=%s links=%s",
        title[:80],
        len(text),
        len(chunks),
        len(parser.links),
    )

    return ExtractedDocument(
        title=title[:300],
        publisher=publisher[:200],
        source_type=classify_source(fetch.final_url, title),
        publication_date=parser.meta.get("article:published_time") or None,
        text=text,
        chunks=chunks,
        links=parser.links[:200],
        metadata={"extractor": "html", "retrieved_at": fetch.retrieved_at},
    )


def extract_pdf(fetch: FetchResult) -> ExtractedDocument:
    """Extract page text and metadata from a PDF fetch result."""
    logger.info("Extracting PDF from %s", fetch.final_url)

    try:
        from pypdf import PdfReader
    except Exception as exc:  # noqa: BLE001
        logger.exception("PDF extraction requires pypdf but import failed")
        raise RuntimeError("PDF extraction requires pypdf") from exc

    reader = PdfReader(io.BytesIO(fetch.body))
    title = ""

    if reader.metadata:
        title = str(reader.metadata.title or "")

    pages: list[str] = []

    for index, page in enumerate(reader.pages, start=1):
        page_text = page.extract_text() or ""

        if page_text.strip():
            pages.append(f"[page {index}] {page_text}")

    text = "\n".join(pages)
    chunks = chunk_text(text)

    logger.info("Extracted PDF pages=%s text_chars=%s chunks=%s", len(reader.pages), len(text), len(chunks))

    return ExtractedDocument(
        title=title or Path(fetch.final_url).name,
        publisher="",
        source_type=classify_source(fetch.final_url, title),
        publication_date=None,
        text=text,
        chunks=chunks,
        links=[],
        metadata={"extractor": "pdf", "pages": len(reader.pages), "retrieved_at": fetch.retrieved_at},
    )


def extract_csv(fetch: FetchResult) -> ExtractedDocument:
    """Extract rows from CSV content into pipe-delimited text."""
    logger.info("Extracting CSV from %s", fetch.final_url)

    content = fetch.body.decode("utf-8", errors="replace")
    rows = list(csv.reader(io.StringIO(content)))
    lines = [" | ".join(cell.strip() for cell in row) for row in rows]
    text = "\n".join(lines)
    chunks = chunk_text(text)

    logger.info("Extracted CSV rows=%s text_chars=%s chunks=%s", len(rows), len(text), len(chunks))

    return ExtractedDocument(
        title=Path(fetch.final_url).name,
        publisher="",
        source_type=classify_source(fetch.final_url),
        publication_date=None,
        text=text,
        chunks=chunks,
        links=[],
        metadata={"extractor": "csv", "rows": len(rows), "retrieved_at": fetch.retrieved_at},
    )


def extract_xlsx(fetch: FetchResult) -> ExtractedDocument:
    """Extract worksheet rows from an XLSX file into normalized text."""
    logger.info("Extracting XLSX from %s", fetch.final_url)

    try:
        from openpyxl import load_workbook
    except Exception as exc:  # noqa: BLE001
        logger.exception("XLSX extraction requires openpyxl but import failed")
        raise RuntimeError("XLSX extraction requires openpyxl") from exc

    workbook = load_workbook(io.BytesIO(fetch.body), read_only=True, data_only=True)
    lines: list[str] = []

    for sheet in workbook.worksheets:
        lines.append(f"[sheet {sheet.title}]")

        for row in sheet.iter_rows(values_only=True):
            values = ["" if value is None else str(value) for value in row]

            if any(value.strip() for value in values):
                lines.append(" | ".join(values))

    text = "\n".join(lines)
    chunks = chunk_text(text)

    logger.info("Extracted XLSX sheets=%s text_chars=%s chunks=%s", len(workbook.sheetnames), len(text), len(chunks))

    return ExtractedDocument(
        title=Path(fetch.final_url).name,
        publisher="",
        source_type=classify_source(fetch.final_url),
        publication_date=None,
        text=text,
        chunks=chunks,
        links=[],
        metadata={"extractor": "xlsx", "sheets": workbook.sheetnames, "retrieved_at": fetch.retrieved_at},
    )


def extract_docx(fetch: FetchResult) -> ExtractedDocument:
    """Extract document XML text from a DOCX file."""
    logger.info("Extracting DOCX from %s", fetch.final_url)

    text_parts: list[str] = []

    with zipfile.ZipFile(io.BytesIO(fetch.body)) as archive:
        xml = archive.read("word/document.xml").decode("utf-8", errors="replace")

    text = re.sub(r"<[^>]+>", " ", xml)
    text = re.sub(r"\s+", " ", text).strip()
    text_parts.append(text)

    joined_text = " ".join(text_parts)
    chunks = chunk_text(joined_text)

    logger.info("Extracted DOCX text_chars=%s chunks=%s", len(joined_text), len(chunks))

    return ExtractedDocument(
        title=Path(fetch.final_url).name,
        publisher="",
        source_type=classify_source(fetch.final_url),
        publication_date=None,
        text=joined_text,
        chunks=chunks,
        links=[],
        metadata={"extractor": "docx", "retrieved_at": fetch.retrieved_at},
    )


def extract_document(fetch: FetchResult) -> tuple[str, ExtractedDocument]:
    """Choose the extractor for a fetch result based on content type and final URL."""
    lowered = f"{fetch.content_type} {fetch.final_url}".lower()

    if "pdf" in lowered or lowered.endswith(".pdf"):
        logger.debug("Selected PDF extractor for %s", fetch.final_url)
        return "pdf", extract_pdf(fetch)

    if "csv" in lowered or lowered.endswith(".csv"):
        logger.debug("Selected CSV extractor for %s", fetch.final_url)
        return "csv", extract_csv(fetch)

    if "spreadsheet" in lowered or lowered.endswith(".xlsx"):
        logger.debug("Selected XLSX extractor for %s", fetch.final_url)
        return "xlsx", extract_xlsx(fetch)

    if "wordprocessingml" in lowered or lowered.endswith(".docx"):
        logger.debug("Selected DOCX extractor for %s", fetch.final_url)
        return "docx", extract_docx(fetch)

    logger.debug("Selected HTML extractor for %s", fetch.final_url)

    return "html", extract_html(fetch)
